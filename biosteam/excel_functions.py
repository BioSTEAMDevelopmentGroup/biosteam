# -*- coding: utf-8 -*-
"""
Created on Wed Sep 27 18:02:09 2017

@author: Yoel Rene Cortes-Pena
"""
from . import np
from openpyxl import Workbook, load_workbook

# TODO: Must comment!

# %% Macking excel file


def create_excel(ID, sheets):
    """
    Create a new excel file.  'ID' is the file ID and 'sheets' is a list of sheet titles.
    Returns the workbook of the created excel file
    """
    if isinstance(sheets, str):
        sheets = [sheets]
    wb = Workbook()
    ws = wb.active
    ws.title = sheets[0]
    l = len(sheets)

    if l > 1:
        for x in range(1, l):
            ws = wb.create_sheet(sheets[x])

    wb.save(ID)
    return wb

# %% Get values as arrays


def array(ws, row_interval=None, col_interval=None):
    # Returns an array of the xml intervals
    if row_interval == None:
        row_interval = (0, len(ws.rows))
    if col_interval == None:
        col_interval = (0, len(ws.columns))
    min_row, max_row = row_interval
    min_col, max_col = col_interval
    M = min_row - max_row
    N = min_col - max_col
    MN = np.array(M, N)
    for m in range(M):
        for n in range(N):
            MN[m, n] = ws.cell(row=m, col=n)
    return MN


def col_arrays(ws, min_row=None, max_row=None):
    # Returns an list of column arrays
    out = []
    for col in ws.iter_cols(min_row=min_row, max_row=max_row):
        l = len(col)
        arr = np.zeros(l)
        for i in range(l):
            arr[i] = col[i].value
        out.append(arr)
    return out


def row_arrays(ws, min_col=None, max_col=None):
    # Returns an list of row arrays
    out = []
    for row in ws.iter_rows(min_col=min_col, max_col=max_col):
        l = len(row)
        arr = np.zeros(l)
        for i in range(l):
            arr[i] = row[i].value
        out.append(arr)
    return out
